import openpyxl
import pandas as pd
import time

if __name__=='__main__':

    t0=time.process_time()
    #Open Workbook
    wb = openpyxl.load_workbook("LoughranMcDonald_MasterDictionary_2018.xlsx")
    ws = wb.active
    print('Workbook Opened')

    #Create 2 lists with positive and negative words
    pos_list, neg_list = [], []
    for r in range(2, ws.max_row + 1):
        word = ws.cell(row=r, column=1).value
        if ws.cell(row=r, column=8).value != 0:
            neg_list.append(word)
        if ws.cell(row=r, column=9).value != 0:
            pos_list.append(word)
    print('Lists Created')

    #Convert into a 1 column dataframe
    neg_df = pd.DataFrame({'word': neg_list})
    pos_df = pd.DataFrame({'word': pos_list})
    print('Dataframe created')

    #Save dataframes as .xlsx files
    neg_writer = pd.ExcelWriter("negative_dictionary.xlsx", engine='openpyxl')
    pos_writer = pd.ExcelWriter("positive_dictionary.xlsx", engine='openpyxl')

    neg_df.to_excel(neg_writer, index=False)
    pos_df.to_excel(pos_writer, index=False)
    print('Data put into .xlsx workbooks')

    #Save the files
    pos_writer.save()
    neg_writer.save()

    print('Workbooks saved')
    print('Execution Time: '+str(time.process_time()-t0))