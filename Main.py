import regex as re
from urllib.request import Request, urlopen
import openpyxl
import encodings.idna
from nltk.tokenize import sent_tokenize
from nltk.tokenize import word_tokenize
from bs4 import BeautifulSoup as bs
import time

#global variables - all names self_explanatory
global stopwords,positive_words,negative_words,uncertainty_words,constraining_words, roman_numerals
global positive_count,negative_count,constraining_count,uncertainty_count,complex_count,sentence_count,word_count,constraining_whole_count

# An operation to force https requests in case they are forbidden on the first attempt
def openfile(request):
    try:
        file = urlopen(request).read()
    except:
        file = openfile(request)
    return file


# Sets the value of the global variables when part of url is input
def getGlobalVariables(string):
    global stopwords, positive_words, negative_words, uncertainty_words, constraining_words, roman_numerals
    global positive_count, negative_count, constraining_count, uncertainty_count, complex_count, sentence_count, word_count, constraining_whole_count

    clean_words = []
    # reset counts
    positive_count, negative_count, uncertainty_count, constraining_count, complex_count, sentence_count, word_count, constraining_whole_count = 0, 0, 0, 0, 0, 0, 0, 0
    url = "https://www.sec.gov/Archives/" + string
    print('url: '+url)
    request = Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    file = openfile(request)
    file_txt = bs(file, 'html.parser').get_text()

    tokenized_sent = sent_tokenize(file_txt)
    sentence_count = len(tokenized_sent)
    tokenized_words = word_tokenize(file_txt)
    alpha_words = [word.lower() for word in tokenized_words if word.isalpha()]
    for word in alpha_words:
        if word.upper() in constraining_words:
            constraining_whole_count+=1
        if word not in stopwords:
            if word not in roman_numerals:
                clean_words.append(word)
                if word.upper() in positive_words: positive_count += 1
                if word.upper() in negative_words: negative_count += 1
                if word.upper() in uncertainty_words: uncertainty_count += 1
                if word.upper() in constraining_words: constraining_count += 1
                if count_syllables(word) > 2: complex_count += 1
    word_count = len(clean_words)


def create_dictionaries():
    global stopwords, positive_words, negative_words, uncertainty_words, constraining_words, roman_numerals

    # Create the roman numeral dict upto 30
    roman_numerals = {'i': 0, 'ii': 0, 'iii': 0, 'iv': 0, 'v': 0,
                      'vi': 0, 'vii': 0, 'viii': 0, 'ix': 0, 'x': 0,
                      'xi': 0, 'xii': 0, 'xiii': 0, 'xiv': 0, 'xv': 0,
                      'xvi': 0, 'xvii': 0, 'xviii': 0, 'xix': 0, 'xx': 0,
                      'xxi': 0, 'xxii': 0, 'xxiii': 0, 'xxiv': 0, 'xxv': 0,
                      'xxvi': 0, 'xxvii': 0, 'xxviii': 0, 'xxix': 0, 'xxx': 0}

    # Getting dictionary of stopwords from the given txt file
    sw_file = open("StopWords_GenericLong.txt")
    for line in sw_file:
        line1 = str(line)
        line1 = line1.strip("\n")
        stopwords.update({line1: 0})

    # using the 4 excel workbooks to create dictionary of positive, negative, uncertainty and constraining words
    pos_wb = openpyxl.load_workbook("positive_dictionary.xlsx")
    neg_wb = openpyxl.load_workbook("negative_dictionary.xlsx")
    con_wb = openpyxl.load_workbook("constraining_dictionary.xlsx")
    uncer_wb = openpyxl.load_workbook("uncertainty_dictionary.xlsx")

    # wb -> workbook, ws -> worksheet
    pos_ws = pos_wb.active
    neg_ws = neg_wb.active
    con_ws = con_wb.active
    uncer_ws = uncer_wb.active

    # Code to read first columnn of every row and update corresponding dictionary
    for rownum in range(2, pos_ws.max_row + 1):
        word = str(pos_ws.cell(row=rownum, column=1).value)
        positive_words.update({word: 0})

    for rownum in range(2, neg_ws.max_row + 1):
        word = str(neg_ws.cell(row=rownum, column=1).value)
        negative_words.update({word: 0})

    for rownum in range(2, con_ws.max_row + 1):
        word = str(con_ws.cell(row=rownum, column=1).value)
        constraining_words.update({word: 0})

    for rownum in range(2, uncer_ws.max_row + 1):
        word = str(uncer_ws.cell(row=rownum, column=1).value)
        uncertainty_words.update({word: 0})

    # Closing workbook after use
    pos_wb.close()
    neg_wb.close()
    con_wb.close()
    uncer_wb.close()

#A great function made through a little help of the internet to count Syllables with accuracy of 0.90 ie. 90%:-

#number of syllables +=1 whenever we encounter a vowel in a word
VOWEL_RUNS = re.compile("[aeiouy]+", flags=re.I)

#Exceptions that subtract number of syllables at occurance of a vowel
EXCEPTIONS = re.compile(
    "[^aeiou]e[sd]?$|"
    + "[^e]ely$",
    flags=re.I
)

#Exceptions that add a syllable to a word
ADDITIONAL = re.compile(
    "[^aeioulr][lr]e[sd]?$|[csgz]es$|[td]ed$|"
    + ".y[aeiou]|ia(?!n$)|eo|ism$|[^aeiou]ire$|[^gq]ua",
    flags=re.I
)

#counts number of syllables in a word
def count_syllables(word):
    vowel_runs = len(VOWEL_RUNS.findall(word))
    exceptions = len(EXCEPTIONS.findall(word))
    additional = len(ADDITIONAL.findall(word))
    return max(1, vowel_runs - exceptions + additional)

def textual_analysis(r):
    global stopwords, positive_words, negative_words, uncertainty_words, constraining_words, roman_numerals
    global positive_count, negative_count, constraining_count, uncertainty_count, complex_count, sentence_count, word_count, constraining_whole_count
    # All textual analysis including storage of final output in the desired location will take place here

    # Get first six columns exactly as they were in the cik file
    cik = str(ws_cik.cell(row=r, column=1).value)
    coname = str(ws_cik.cell(row=r, column=2).value)
    fyrmo = str(ws_cik.cell(row=r, column=3).value)
    fdate = str(ws_cik.cell(row=r, column=4).value)
    form = str(ws_cik.cell(row=r, column=5).value)
    secfname = str(ws_cik.cell(row=r, column=6).value)

    # Put these column values into the output file
    ws_output.cell(row=r, column=1, value=cik)
    ws_output.cell(row=r, column=2, value=coname)
    ws_output.cell(row=r, column=3, value=fyrmo)
    ws_output.cell(row=r, column=4, value=fdate)
    ws_output.cell(row=r, column=5, value=form)
    ws_output.cell(row=r, column=6, value=secfname)

    # col 7,8
    ws_output.cell(row=r, column=7, value=positive_count)
    ws_output.cell(row=r, column=8, value=negative_count)

    # col 9,10
    polarity_score = 0.000001 + (positive_count - negative_count) / (positive_count + negative_count)
    avg_sent_len = word_count / sentence_count

    ws_output.cell(row=r, column=9, value=polarity_score)
    ws_output.cell(row=r, column=10, value=avg_sent_len)

    # col 11,12
    perc_complex_words = complex_count / word_count
    fog_index = 0.4 * (avg_sent_len + perc_complex_words)

    ws_output.cell(row=r, column=11, value=perc_complex_words)
    ws_output.cell(row=r, column=12, value=fog_index)

    # col 13,14
    ws_output.cell(row=r, column=13, value=complex_count)
    ws_output.cell(row=r, column=14, value=word_count)

    # col 15,16
    ws_output.cell(row=r, column=15, value=uncertainty_count)
    ws_output.cell(row=r, column=16, value=constraining_count)

    # col 17,18,19,20
    pos_prop = positive_count / word_count
    neg_prop = negative_count / word_count
    uncer_prop = uncertainty_count / word_count
    con_prop = constraining_count / word_count

    ws_output.cell(row=r, column=17, value=pos_prop)
    ws_output.cell(row=r, column=18, value=neg_prop)
    ws_output.cell(row=r, column=19, value=uncer_prop)
    ws_output.cell(row=r, column=20, value=con_prop)

    # col 21
    ws_output.cell(row=r, column=21, value=constraining_whole_count)

if __name__ == '__main__':
    global stopwords, positive_words, negative_words, uncertainty_words, constraining_words, roman_numerals
    global positive_count, negative_count, constraining_count, uncertainty_count, complex_count, sentence_count, word_count, constraining_whole_count

    # initilizing global variables
    stopwords, positive_words, negative_words, uncertainty_words, constraining_words, roman_numerals = dict(), dict(), dict(), dict(), dict(), dict()
    positive_count, negative_count, constraining_count, uncertainty_count, complex_count, sentence_count, word_count, constraining_whole_count = 0, 0, 0, 0, 0, 0, 0, 0

    # Creating the desired dictionaries
    create_dictionaries()

    # Loading input and output files
    wb_cik = openpyxl.load_workbook('cik_list.xlsx')
    ws_cik = wb_cik.active
    wb_output = openpyxl.load_workbook('Output Data.xlsx')
    ws_output = wb_output.active

    # Iterative code to get url from every line and run the functions to get desired result
    print('Program Started')
    t0=time.process_time()
    for rownum in range(2, ws_cik.max_row + 1):
        t1=time.process_time()
        string = ws_cik.cell(row=rownum, column=6).value
        global_variables = getGlobalVariables(str(string))
        textual_analysis(rownum)
        print('Time taken to process this url: '+str(time.process_time()-t1))

    #Saving data and closing workbook
    wb_output.save('Output Data.xlsx')
    wb_cik.close()
    wb_output.close()
    print("Total process time: "+str(time.process_time()-t0))
    print('Program End')